"""
Comando Django para importar el Excel "errores Dimarsa.xlsx" a la base SQLite.

Uso:
    python manage.py importar_excel ruta/al/archivo.xlsx
    python manage.py importar_excel ruta/al/archivo.xlsx --reset
"""
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import IntegrityError

from errores.dominios.catalogos import Frecuencia, Prueba
from errores.dominios.errores import Error, ErrorPrueba, ErrorSolucion
from errores.dominios.plataforma import Hardware, HardwareSistema, Modulo, Sistema
from errores.dominios.soluciones import Solucion, SolucionAsignacion


def s(value):
    """Convierte un valor de celda a string limpio o cadena vacía."""
    if value is None:
        return ''
    return str(value).strip()


def key(value):
    return s(value).casefold()


class Command(BaseCommand):
    help = 'Importa el Excel "errores Dimarsa.xlsx" a la base de datos.'

    def add_arguments(self, parser):
        parser.add_argument('excel_path', type=str, help='Ruta al archivo .xlsx')
        parser.add_argument(
            '--reset',
            action='store_true',
            help='Elimina los errores existentes antes de importar.',
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            raise CommandError("Falta openpyxl. Instala con: pip install openpyxl")

        excel_path = Path(options['excel_path'])
        if not excel_path.exists():
            raise CommandError(f'No existe el archivo: {excel_path}')

        self.stdout.write(self.style.NOTICE(f'Leyendo {excel_path}...'))
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        if options['reset']:
            self.stdout.write(self.style.WARNING('Eliminando registros existentes...'))
            SolucionAsignacion.objects.all().delete()
            ErrorSolucion.objects.all().delete()
            ErrorPrueba.objects.all().delete()
            Error.objects.all().delete()
            HardwareSistema.objects.all().delete()
            Modulo.objects.all().delete()
            Prueba.objects.all().delete()
            Hardware.objects.all().delete()
            Sistema.objects.all().delete()
            Frecuencia.objects.all().delete()
            Solucion.objects.all().delete()

        self._iniciar_caches()
        self._importar_pruebas_tipicas(wb)
        self._importar_hardware(wb)
        self._importar_sistemas(wb)
        self._importar_hardware_sistemas(wb)
        self._importar_frecuencias(wb)
        self._importar_soluciones(wb)
        self._importar_errores(wb)

        self.stdout.write(self.style.SUCCESS('¡Importación completada!'))
        self.stdout.write(f'  Hardware: {Hardware.objects.count()}')
        self.stdout.write(f'  Sistemas: {Sistema.objects.count()}')
        self.stdout.write(f'  Modulos: {Modulo.objects.count()}')
        self.stdout.write(f'  Frecuencias: {Frecuencia.objects.count()}')
        self.stdout.write(f'  Soluciones: {Solucion.objects.count()}')
        self.stdout.write(f'  Pruebas: {Prueba.objects.count()}')
        self.stdout.write(f'  Errores: {Error.objects.count()}')

    # ------------------------------------------------------------------ #
    def _iniciar_caches(self):
        self.hardware_cache = {key(obj.nombre): obj for obj in Hardware.objects.all()}
        self.sistema_cache = {key(obj.nombre): obj for obj in Sistema.objects.all()}
        self.frecuencia_cache = {key(obj.nivel): obj for obj in Frecuencia.objects.all()}
        self.solucion_cache = {key(obj.nombre): obj for obj in Solucion.objects.all()}
        self.prueba_cache = {key(obj.descripcion): obj for obj in Prueba.objects.all()}
        self.modulo_cache = {(obj.sistema_id, key(obj.nombre)): obj for obj in Modulo.objects.all()}
        self.hw_sistema_cache = set(HardwareSistema.objects.values_list('hardware_id', 'sistema_id'))

    def _get_hardware(self, nombre, so=''):
        nombre = nombre[:120]
        if not nombre:
            return None
        cache_key = key(nombre)
        obj = self.hardware_cache.get(cache_key)
        if obj is None:
            try:
                obj = Hardware.objects.create(nombre=nombre, so=so[:160])
            except IntegrityError:
                obj = Hardware.objects.get(nombre=nombre)
            self.hardware_cache[cache_key] = obj
        return obj

    def _get_sistema(self, nombre):
        nombre = nombre[:160]
        if not nombre:
            return None
        cache_key = key(nombre)
        obj = self.sistema_cache.get(cache_key)
        if obj is None:
            try:
                obj = Sistema.objects.create(nombre=nombre)
            except IntegrityError:
                obj = Sistema.objects.get(nombre=nombre)
            self.sistema_cache[cache_key] = obj
        return obj

    def _get_frecuencia(self, nivel):
        nivel = nivel[:80]
        if not nivel:
            return None
        cache_key = key(nivel)
        obj = self.frecuencia_cache.get(cache_key)
        if obj is None:
            try:
                obj = Frecuencia.objects.create(nivel=nivel)
            except IntegrityError:
                obj = Frecuencia.objects.get(nivel=nivel)
            self.frecuencia_cache[cache_key] = obj
        return obj

    def _get_solucion(self, nombre, pasos='', anexos=''):
        nombre = nombre[:220]
        if not nombre:
            return None
        cache_key = key(nombre)
        obj = self.solucion_cache.get(cache_key)
        if obj is None:
            try:
                obj = Solucion.objects.create(nombre=nombre, pasos=pasos, anexos=anexos)
            except IntegrityError:
                obj = Solucion.objects.get(nombre=nombre)
            self.solucion_cache[cache_key] = obj
        return obj

    def _get_prueba(self, descripcion):
        descripcion = descripcion[:255]
        if not descripcion:
            return None
        cache_key = key(descripcion)
        obj = self.prueba_cache.get(cache_key)
        if obj is None:
            try:
                obj = Prueba.objects.create(descripcion=descripcion)
            except IntegrityError:
                obj = Prueba.objects.get(descripcion=descripcion)
            self.prueba_cache[cache_key] = obj
        return obj

    def _get_modulo(self, sistema, nombre):
        nombre = nombre[:180]
        if not sistema or not nombre:
            return None
        cache_key = (sistema.id, key(nombre))
        obj = self.modulo_cache.get(cache_key)
        if obj is None:
            try:
                obj = Modulo.objects.create(sistema=sistema, nombre=nombre)
            except IntegrityError:
                obj = Modulo.objects.get(sistema=sistema, nombre=nombre)
            self.modulo_cache[cache_key] = obj
        return obj

    def _relacionar_hardware_sistema(self, hardware, sistema):
        if not hardware or not sistema:
            return
        key = (hardware.id, sistema.id)
        if key not in self.hw_sistema_cache:
            try:
                HardwareSistema.objects.create(hardware=hardware, sistema=sistema)
            except IntegrityError:
                pass
            self.hw_sistema_cache.add(key)

    def _importar_pruebas_tipicas(self, wb):
        if 'Típico' not in wb.sheetnames:
            return
        ws = wb['Típico']
        for row in ws.iter_rows(min_row=2, values_only=True):
            self._get_prueba(s(row[1]))

    def _importar_hardware(self, wb):
        if 'Hardware' not in wb.sheetnames:
            return
        ws = wb['Hardware']
        for row in ws.iter_rows(min_row=3, values_only=True):
            # (None, ID, Tipo, SO)
            tipo = s(row[2])
            so = s(row[3])
            if not tipo:
                continue
            self._get_hardware(tipo, so)

    def _importar_sistemas(self, wb):
        if 'Sistemas' not in wb.sheetnames:
            return
        ws = wb['Sistemas']
        for row in ws.iter_rows(min_row=3, values_only=True):
            # (None, ID, Nombre, Sección)
            nombre = s(row[2])
            seccion = s(row[3])
            if not nombre:
                continue
            sistema = self._get_sistema(nombre)
            if seccion:
                self._get_modulo(sistema, seccion)

    def _importar_hardware_sistemas(self, wb):
        if 'Hard-Sist' not in wb.sheetnames:
            return
        ws = wb['Hard-Sist']
        hardware_actual = None
        for row in ws.iter_rows(min_row=3, values_only=True):
            tipo = s(row[2])
            sistema_nombre = s(row[4])
            if tipo:
                hardware_actual = self._get_hardware(tipo)
            if not hardware_actual or not sistema_nombre:
                continue
            sistema = self._get_sistema(sistema_nombre)
            self._relacionar_hardware_sistema(hardware_actual, sistema)

    def _importar_frecuencias(self, wb):
        if 'Frecuencia' not in wb.sheetnames:
            return
        ws = wb['Frecuencia']
        for row in ws.iter_rows(min_row=3, values_only=True):
            nivel = s(row[2])
            if not nivel:
                continue
            self._get_frecuencia(nivel)

    def _importar_soluciones(self, wb):
        if 'Soluciones' not in wb.sheetnames:
            return
        ws = wb['Soluciones']
        for row in ws.iter_rows(min_row=3, values_only=True):
            # (None, ID, Solución, Pasos, Video)
            nombre = s(row[2])
            pasos = s(row[3])
            video = s(row[4])
            if not nombre:
                continue
            self._get_solucion(nombre, pasos, video)

    def _importar_errores(self, wb):
        if 'Errores' not in wb.sheetnames:
            raise CommandError("La hoja 'Errores' no existe en el archivo")
        ws = wb['Errores']

        filas = []
        hardware_nombres = set()
        sistema_nombres = set()
        frecuencia_niveles = set()
        prueba_descripciones = set()
        solucion_textos = set()
        modulo_claves = set()

        for row in ws.iter_rows(min_row=3, values_only=True):
            # Columnas: A=ID_Error, B=Descripción, C=Imagen, D=Usuario,
            # E=Hardware, F=Sistema, G=Sección, H=Pruebas, I=Causa,
            # J=Solución, K=Frecuencia, L=Comentarios
            descripcion = s(row[1])
            if not descripcion:
                continue

            solucion_texto = s(row[9])
            hw_nombre = s(row[4])[:120]
            sys_nombre = s(row[5])[:160]
            seccion = s(row[6])[:180]
            pruebas = s(row[7])[:255]
            freq_nivel = s(row[10])[:80]

            if hw_nombre:
                hardware_nombres.add(hw_nombre)
            if sys_nombre:
                sistema_nombres.add(sys_nombre)
            if freq_nivel:
                frecuencia_niveles.add(freq_nivel)
            if pruebas:
                prueba_descripciones.add(pruebas)
            if solucion_texto:
                solucion_textos.add(solucion_texto)
            if sys_nombre and seccion:
                modulo_claves.add((sys_nombre, seccion))

            filas.append({
                'descripcion': descripcion,
                'imagen_error': s(row[2]),
                'usuario': s(row[3]),
                'hardware': hw_nombre,
                'sistema': sys_nombre,
                'modulo': seccion,
                'causa': s(row[8]),
                'solucion': solucion_texto,
                'frecuencia': freq_nivel,
                'comentarios': s(row[11]),
            })

        self._bulk_hardware(hardware_nombres)
        self._bulk_sistemas(sistema_nombres)
        self._bulk_frecuencias(frecuencia_niveles)
        self._bulk_pruebas(prueba_descripciones)
        self._bulk_soluciones(solucion_textos)
        self._bulk_modulos(modulo_claves)
        self._bulk_hardware_sistemas({
            (fila['hardware'], fila['sistema'])
            for fila in filas
            if fila['hardware'] and fila['sistema']
        })

        errores = []
        for fila in filas:
            sistema = self.sistema_cache.get(key(fila['sistema']))
            modulo = None
            if sistema and fila['modulo']:
                modulo = self.modulo_cache.get((sistema.id, key(fila['modulo'])))
            errores.append(Error(
                descripcion=fila['descripcion'],
                imagen_error=fila['imagen_error'],
                usuario=fila['usuario'],
                hardware=self.hardware_cache.get(key(fila['hardware'])),
                sistema=sistema,
                modulo=modulo,
                causa=fila['causa'],
                frecuencia=self.frecuencia_cache.get(key(fila['frecuencia'])),
                comentarios=fila['comentarios'],
            ))

        errores_creados = Error.objects.bulk_create(errores, batch_size=200)
        error_soluciones = []
        for error, fila in zip(errores_creados, filas):
            solucion = self.solucion_cache.get(key(fila['solucion'][:220]))
            if solucion:
                error_soluciones.append(ErrorSolucion(error=error, solucion=solucion))
        if error_soluciones:
            ErrorSolucion.objects.bulk_create(error_soluciones, ignore_conflicts=True, batch_size=200)
        self.stdout.write(self.style.SUCCESS(f'  {len(errores)} errores creados'))

    def _bulk_hardware(self, nombres):
        nuevos = [
            Hardware(nombre=nombre)
            for nombre in nombres
            if key(nombre) not in self.hardware_cache
        ]
        if nuevos:
            Hardware.objects.bulk_create(nuevos, ignore_conflicts=True, batch_size=200)
            self.hardware_cache = {key(obj.nombre): obj for obj in Hardware.objects.all()}

    def _bulk_sistemas(self, nombres):
        nuevos = [
            Sistema(nombre=nombre)
            for nombre in nombres
            if key(nombre) not in self.sistema_cache
        ]
        if nuevos:
            Sistema.objects.bulk_create(nuevos, ignore_conflicts=True, batch_size=200)
            self.sistema_cache = {key(obj.nombre): obj for obj in Sistema.objects.all()}

    def _bulk_frecuencias(self, niveles):
        nuevos = [
            Frecuencia(nivel=nivel)
            for nivel in niveles
            if key(nivel) not in self.frecuencia_cache
        ]
        if nuevos:
            Frecuencia.objects.bulk_create(nuevos, ignore_conflicts=True, batch_size=200)
            self.frecuencia_cache = {key(obj.nivel): obj for obj in Frecuencia.objects.all()}

    def _bulk_pruebas(self, descripciones):
        nuevos = [
            Prueba(descripcion=descripcion)
            for descripcion in descripciones
            if key(descripcion) not in self.prueba_cache
        ]
        if nuevos:
            Prueba.objects.bulk_create(nuevos, ignore_conflicts=True, batch_size=200)
            self.prueba_cache = {key(obj.descripcion): obj for obj in Prueba.objects.all()}

    def _bulk_soluciones(self, textos):
        nuevos = []
        for texto in textos:
            nombre = texto[:220]
            if key(nombre) not in self.solucion_cache:
                nuevos.append(Solucion(nombre=nombre, pasos=texto))
        if nuevos:
            Solucion.objects.bulk_create(nuevos, ignore_conflicts=True, batch_size=200)
            self.solucion_cache = {key(obj.nombre): obj for obj in Solucion.objects.all()}

    def _bulk_modulos(self, claves):
        nuevos = []
        for sistema_nombre, modulo_nombre in claves:
            sistema = self.sistema_cache.get(key(sistema_nombre))
            if sistema and (sistema.id, key(modulo_nombre)) not in self.modulo_cache:
                nuevos.append(Modulo(sistema=sistema, nombre=modulo_nombre))
        if nuevos:
            Modulo.objects.bulk_create(nuevos, ignore_conflicts=True, batch_size=200)
            self.modulo_cache = {(obj.sistema_id, key(obj.nombre)): obj for obj in Modulo.objects.all()}

    def _bulk_hardware_sistemas(self, claves):
        nuevos = []
        for hardware_nombre, sistema_nombre in claves:
            hardware = self.hardware_cache.get(key(hardware_nombre))
            sistema = self.sistema_cache.get(key(sistema_nombre))
            if hardware and sistema and (hardware.id, sistema.id) not in self.hw_sistema_cache:
                nuevos.append(HardwareSistema(hardware=hardware, sistema=sistema))
        if nuevos:
            HardwareSistema.objects.bulk_create(nuevos, ignore_conflicts=True, batch_size=200)
            self.hw_sistema_cache = set(HardwareSistema.objects.values_list('hardware_id', 'sistema_id'))
